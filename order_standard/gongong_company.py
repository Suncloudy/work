from openpyxl  import load_workbook
from openpyxl  import Workbook
import os,re,json

import pymysql
from DBUtils.PooledDB import PooledDB 


MYSQL_HOST = '43.247.184.94'
USER = 'nice'
PASSWORD = 'Niceee@2020'
DB = 'gsb_report'
PORT = 7800

class DBpool(object):
    def __init__(self):
        # 5为连接池里的最少连接
        self.pool = PooledDB(pymysql, 5, host=MYSQL_HOST, user=USER, passwd=PASSWORD, db=DB, port=PORT)
    
    def insertData(self,sql,a):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.executemany(sql,a)
            conn.commit()
        except Exception as e:
            print(e)
        cursor.close()
        conn.close()


    def getData(self,sql):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.execute(sql)
            sss = cursor.fetchall()
        except Exception as e:
            sss=None
            print(e)
        cursor.close()
        conn.close()

        return sss


def ReadCompanyExcel(fileName,sheetName):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb[sheetName]
    max_row = sheet.max_row
    resultlist =[]

    alist = []
    for r in range(2,max_row+1):
     # 数字是代表公司名在表格的哪一列，必须设置
        a = sheet.cell(r,1).value
        alist.append(a)
    rwb.close()
    return alist


def WriteManyColExcel(fileName,sheetName,data):            #多列数据插入
    rwb = load_workbook(filename = fileName+'.xlsx')
    allSheet = rwb.sheetnames       # 获取所有的sheet
    if(sheetName not in allSheet):
        rwb.create_sheet(sheetName)
    sheet = rwb[sheetName]
    for i,val in enumerate(data):
        for j,v in enumerate(val):
            sheet.cell(i+1,j+1,v)

    rwb.save(fileName+'.xlsx')
    rwb.close()


def WriteOneColExcel(fileName,sheetName,data):            #单列数据插入
    rwb = load_workbook(filename = fileName+'.xlsx')
    allSheet = rwb.sheetnames       # 获取所有的sheet
    if(sheetName not in allSheet):
        rwb.create_sheet(sheetName)

    sheet = rwb[sheetName]
    for i,val in enumerate(data):
        sheet.cell(i+1,1,val)

    rwb.save(fileName+'.xlsx')
    rwb.close()



# 获取两个数据的交集
def getNewArr(arr1,arr2):
    return [val for val in arr1 if val in arr2]
    
if __name__ == "__main__":


    # ali_companies = ReadCompanyExcel("阿里云渠道-含工商数据","Sheet4")
    ali_companies = ReadCompanyExcel("company_new1","Sheet")
    print(len(ali_companies))

    pro_companies = ReadCompanyExcel("推荐产品过滤结果0505","Sheet1")
    print(len(pro_companies))

    company_list = getNewArr(ali_companies,pro_companies)

    WriteOneColExcel("company_new","Sheet3",company_list)

