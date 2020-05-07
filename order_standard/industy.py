
from openpyxl  import load_workbook
from openpyxl  import Workbook
import os,re,json

import pymysql
from DBUtils.PooledDB import PooledDB 


MYSQL_HOST = '43.247.184.94'
USER = 'nice'
PASSWORD = 'Niceee@2020'
DB = 'gsb_report'
PORT = 7800

class DBpool(object):
    def __init__(self):
        # 5为连接池里的最少连接
        self.pool = PooledDB(pymysql, 5, host=MYSQL_HOST, user=USER, passwd=PASSWORD, db=DB, port=PORT)
    
    def insertData(self,sql,a):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.executemany(sql,a)
            conn.commit()
        except Exception as e:
            print(e)
        cursor.close()
        conn.close()


    def getData(self,sql):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.execute(sql)
            sss = cursor.fetchall()
        except Exception as e:
            sss=None
            print(e)
        cursor.close()
        conn.close()

        return sss


def ReadCompanyExcel(fileName):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet5"]
    max_row = sheet.max_row
    resultlist =[]

    adict = {}
    for r in range(1,max_row+1):
     # 数字是代表公司名在表格的哪一列，必须设置
        a = sheet.cell(r,1).value
        b = sheet.cell(r,2).value
        adict[a]=b
    rwb.close()

    return adict
    


def ReadCompanyExcel222(fileName):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet"]
    max_row = sheet.max_row
    resultlist =[]

    abc= []
    for r in range(2,max_row+1):
        
     # 数字是代表公司名在表格的哪一列，必须设置

        c = sheet.cell(r,4).value
        d = sheet.cell(r,2).value
        adict = [d,c] 
        abc.append(adict)
        adict=[]

    rwb.close()

    return abc
    
def ReadCompanyExcel1(fileName,lss):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet4"]
    max_row = sheet.max_row
    resultlist =[]

    print(sss)

    for r in range(2,max_row+1):
     # 数字是代表公司名在表格的哪一列，必须设置
        ax = sheet.cell(r,2).value
        print(ax)
        abc = lss[ax]
        sheet.cell(r,1,abc)
    rwb.save(fileName+'.xlsx')


def readJson():
    current_path = os.path.dirname(os.path.abspath(__file__))
    print(current_path)
    with open(current_path+"\\record.json",'r',encoding="UTF-8") as load_f:
        record = json.load(load_f)
        load_f.close()
        
    # with open(current_path+"\\record2.json",'r',encoding="UTF-8") as load_f2:
    #     record2 = json.load(load_f2)
    #     load_f2.close()
        
    return record["data"]


if __name__ == "__main__":

    # db = DBpool()

    fileName="company_new"
    sss = ReadCompanyExcel222(fileName)

    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet"]
    
    rec = readJson()

    for i, s in enumerate(sss):
        name = s[0]
        ind = s[1]
        print(ind)
        data = rec[ind] if ind !=None and ind!='' else None
        sheet.cell(i+2,6,data)

    rwb.save(fileName+'.xlsx')
    rwb.close()