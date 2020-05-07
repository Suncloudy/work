""" 
    getCompnayByName    根据 公司名称查询并写入到 excel文件中
    getCompnayById      根据 公司ID 查询并写入到 excel文件中

 """


from elasticsearch import Elasticsearch
import time
from openpyxl  import load_workbook
from openpyxl  import Workbook
import os,re

# esip = "39.106.39.121"
# esprot = "8080"
# esuser = "esroot"
# espassword = "esroot"

esip = "43.247.184.94"
esprot = "9200"
esuser = "admines"
espassword = "adminGSB."



# 将公司名称写入到txt文件中
# def writeTxt(company_name):
#     if(os.path.exists())
#     if(type(company_name)==str):


# 创建excel文件
def mkXlsx(file_name):
    wb = Workbook()
    ws = wb.active
    wb.save(file_name)

def qingxi(ustring):
    """全角转半角"""
    if ustring:
        rstring = ""
        for uchar in ustring:
            inside_code = ord(uchar)
            if inside_code == 12288:  # 全角空格直接转换
                inside_code = 32
            elif (inside_code >= 65281 and inside_code <= 65374):  # 全角字符（除空格）根据关系转化
                inside_code -= 65248
            rstring += chr(inside_code)
        for line in list(rstring):
            if line in ['·', '(', ')']:
                return rstring
        # if '·' or '(' or ')' in rstring:
        #    return  rstring
        data = re.sub('([\W,_]+)', '', rstring)
        return data

def ReadCompanyExcel(fileName):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet3"]
    max_row = sheet.max_row
    resultlist =[]

    for r in range(1,max_row+1):
     # 数字是代表公司名在表格的哪一列，必须设置
        result = sheet.cell(r,1).value
        if(result!=None):
            if(type(result)==int or type(result)==float):
                resultlist.append(result)
            else:
                result.replace(" ", "")
                res = qingxi(result)
                resultlist.append(res)
        else:
            resultlist.append('__')

    return resultlist


# 将 company 信息写入到 xlsx 文件中
def writeData(file_name,companies):
    arr1 = [
        'company_id','company_name','base','gsb_city','legal_person','reg_status','reg_capital','reg_unit','estiblish_time',\
            'credit_code','reg_number','org_number','company_type','reg_institute','reg_location','business_scope','from_time',\
                'to_time','approved_time','old_name','company_org_type','gsb_company_cate_1','gsb_company_cate_2','gsb_company_cate_3'
    ]
    rwb = load_workbook(filename = file_name+'.xlsx',)
    #rwb.create_sheet(file_name,0)
    sheet = rwb['Sheet']
    max_row = sheet.max_row
    # max_row = row
    if(max_row==1):
        #print(1)
        for index,field in enumerate(arr1):
            sheet.cell(1,index+1,field)
        for i,company in enumerate(companies):
            #print(i)
            if(company!=[]):
                for index,value in enumerate(company):
                    sheet.cell(i+2,index+1,value)
            else:
                pass
    else:
        #print(2)
        for company in companies:
            for index,value in enumerate(company):
                sheet.cell(max_row+1,index+1,value) 

    rwb.save(file_name+'.xlsx')

def writeXlsxFile(file,companies):       # companies 包括        三个属性：文件名称、字段名称、字段对应的值
    fileName = file +'.xlsx'
    #print(fileName)
    boolean = os.path.exists(fileName)
    if(boolean==True):
        writeData(file,companies)
    else:
        mkXlsx(fileName)
        writeData(file,companies)

def ToDateTime(timestamp):
    if(timestamp!=None):
        timeArray = time.localtime(timestamp)
        otherStyleTime = time.strftime("%Y/%m/%d", timeArray)       #   %Y/%m/%d %H:%M:%S
    else:
        otherStyleTime = ''
    return otherStyleTime

def ifNone(field):
    if(field==None):
        return ''
    else:
        return field

def JsonToArr(esJson):
    index = esJson["_index"]
    company_id = esJson["_id"]
    source = esJson["_source"]

    company_name = ifNone(source["company_name"])
    base = ifNone(source["base"])
    gsb_city = ifNone(source["gsb_city"])
    legal_person = ifNone(source["legal_person"])
    reg_status = ifNone(source["reg_status"])
    reg_unit = ifNone(source["reg_unit"])
    reg_capital = ifNone(source["reg_capital"])
    estiblish_time = ToDateTime(source["estiblish_time"])
    credit_code = ifNone(source["credit_code"])
    reg_number = ifNone(source["reg_number"])
    org_number = ifNone(source["org_number"])
    company_type = ifNone(source["company_type"])
    reg_institute = ifNone(source["reg_institute"])
    reg_location = ifNone(source["reg_location"])
    business_scope = ifNone(source["business_scope"])
    from_time = ToDateTime(source["from_time"])
    to_time = ToDateTime(source["to_time"])
    approved_time = ToDateTime(source["approved_time"])
    old_name = ifNone(source["old_name"])
    company_org_type = ifNone(source["company_org_type"])
    gsb_company_cate_1 = ifNone(source["gsb_company_cate_1"])
    gsb_company_cate_2 = ifNone(source["gsb_company_cate_2"])
    gsb_company_cate_3 = ifNone(source["gsb_company_cate_3"])



    arr2 = [
        company_id,company_name,base,gsb_city,legal_person,reg_status,reg_capital,reg_unit,estiblish_time,credit_code,\
            reg_number,org_number,company_type,reg_institute,reg_location,business_scope,from_time,to_time,approved_time,\
                old_name,company_org_type,gsb_company_cate_1,gsb_company_cate_2,gsb_company_cate_3
    ]
    return arr2

class Doelastic(object):
    def __init__(self):
        self.es = Elasticsearch([esip], http_auth=(esuser, espassword),port=esprot)

    def SearchCompany(self,factor):
        body = {
            "query":{
                "term":{
                    factor[0]:factor[1] 
                }
            }
            #"size":"10000"
        }

        results = self.es.search(index="bigdata_ic_gsb_company_op",body=body)
        # print(results)
        res = results["hits"]["hits"]
        #print(factor[1],'----------------',len(res))
        #print(res)
        companies=[]
        index = "bigdata_ic_gsb_company_op"
        if(res!=[]):
            for re in res:
                company = JsonToArr(re)
                companies.append(company)
            #print(len(companies))
            #writeXlsxFile(index,companies)
             
        else:
            pass
        return companies

    def getCompnayByName(self,company_name):
        factor = ['company_domain',company_name]
        sss = self.SearchCompany(factor)
        return sss

    def getCompnayById(self,company_id):
        factor = ['id',company_id]
        self.SearchCompany(factor)

    def getCompnayByCate(self,cate):
        factor = ['gsb_company_cate_1',cate]
        self.SearchCompany(factor)

    # def MultiFactor(self,factors):
    #     body = {
    #         "query":{
    #             "bool":{
    #                 "must":{
    #                 }
    #             }
    #         }
    #     }



if __name__ == "__main__":
    el = Doelastic()
    # el.getCompnayByCate('教育')
    companies = ReadCompanyExcel('company_new2')

    company_list=[]
    for company in companies:
        print(company)
        ss = el.getCompnayByName(company) 
        #print(ss)
        if(ss!=[]):
            company_list.append(ss[0])
        else:
            company_list.append(ss)

    writeXlsxFile('company_new',company_list)
