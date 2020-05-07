
from openpyxl  import load_workbook
from openpyxl  import Workbook
import os,re

import pymysql
from DBUtils.PooledDB import PooledDB 


MYSQL_HOST = '47.95.76.74'
USER = 'root'
PASSWORD = 'Gongsibao2018'
DB = 'ent_data'
PORT = 3306

class DBpool(object):
    def __init__(self):
        # 5为连接池里的最少连接
        self.pool = PooledDB(pymysql, 5, host=MYSQL_HOST, user=USER, passwd=PASSWORD, db=DB, port=PORT)
    
    def insertData(self,sql,a):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.executemany(sql,a)
            conn.commit()
        except Exception as e:
            print(e)
        cursor.close()
        conn.close()


    def getData(self,sql):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.execute(sql)
            sss = cursor.fetchall()
        except Exception as e:
            sss=None
            print(e)
        cursor.close()
        conn.close()

        return sss


def ReadCompanyExcel(fileName):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet"]
    max_row = sheet.max_row
    resultlist =[]

    adict = {}
    for r in range(1,max_row+1):
     # 数字是代表公司名在表格的哪一列，必须设置
        a = sheet.cell(r,1).value
        b = sheet.cell(r,2).value
        adict[a]=b
    rwb.close()

    return adict
    


def ReadCompanyExcel222(fileName):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet"]
    max_row = sheet.max_row
    resultlist =[]

    abc= []
    for r in range(3,max_row+1):
     # 数字是代表公司名在表格的哪一列，必须设置
        a = sheet.cell(r,2).value
        abc.append(a)
    rwb.close()

    return abc

def ReadCompanyExcel1(fileName,lss):
    rwb = load_workbook(filename = fileName+'.xlsx')
    sheet = rwb["Sheet4"]
    max_row = sheet.max_row
    resultlist =[]

    print(sss)

    for r in range(2,max_row+1):
     # 数字是代表公司名在表格的哪一列，必须设置
        ax = sheet.cell(r,2).value
        print(ax)
        abc = lss[ax]
        sheet.cell(r,1,abc)
    
    rwb.save(fileName+'.xlsx')


if __name__ == "__main__":

    db = DBpool()

    fileName="company_new"
    sss = ReadCompanyExcel222(fileName)

    s1 = []
    s2 = []
    for ss in sss:
        if('(' in ss):
            s1.append(ss)
        else:
            s2.append(ss)


    li = []
    for s in s1:
        sql = "select merge_product_id,merge_product_name,order_add_time,order_price,order_id,order_pkid,product_id,product_name from tj_history_order where account_company_name='%s' and merge_product_id is not null" %(s)
        order_list = ()
        res = db.getData(sql)
        if(res!=None):
            for r in res:
                merge_product_id = r[0]
                merge_product_name =r[1]
                order_add_time = r[2]
                order_price = r[3]
                order_id = r[4]
                order_pkid = r[5]
                product_id = r[6]
                product_name = r[7]
                source = "阿里云"

                order_list = (s,s,merge_product_id,merge_product_name,order_add_time,order_price,source,order_id,order_pkid,product_id,product_name)
                li.append(order_list)
            if(len(li)>200):
                sql2 = "insert into tj_standard_order(standard_company_name,account_company_name,merge_product_id,merge_product_name,order_add_time,order_pay,order_channel,order_id,order_pkid,product_id,product_name)\
                    values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                db.insertData(sql2,li)
                li = []
            else:
                pass
        else:
            s = s.replace("(","（").replace(")","）")
            sql = "select merge_product_id,merge_product_name,order_add_time,order_price,order_id,order_pkid,product_id,product_name from tj_history_order where account_company_name='%s' and merge_product_id is not null" %(s)
            order_list = ()
            res = db.getData(sql)
            if(res!=None):
                for r in res:
                    merge_product_id = r[0]
                    merge_product_name =r[1]
                    order_add_time = r[2]
                    order_price = r[3]
                    order_id = r[4]
                    order_pkid = r[5]
                    product_id = r[6]
                    product_name = r[7]
                    source = "阿里云"

                    order_list = (s,s,merge_product_id,merge_product_name,order_add_time,order_price,source,order_id,order_pkid,product_id,product_name)
                    li.append(order_list)
                if(len(li)>200):
                    sql2 = "insert into tj_standard_order(standard_company_name,account_company_name,merge_product_id,merge_product_name,order_add_time,order_pay,order_channel,order_id,order_pkid,product_id,product_name)\
                    values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    db.insertData(sql2,li)
                    li = []
                else:
                    pass

    sql3 = "insert into tj_standard_order(standard_company_name,account_company_name,merge_product_id,merge_product_name,order_add_time,order_pay,order_channel,order_id,order_pkid,product_id,product_name)\
        values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
    db.insertData(sql3,li)

    li = []
    for s in s2:
        sql1 = "select merge_product_id,merge_product_name,order_add_time,order_price,order_id,order_pkid,product_id,product_name from tj_history_order where account_company_name='%s' and merge_product_id is not null" %(s)
        order_list = ()
        res = db.getData(sql1)
        if(res!=None):
            for r in res:
                merge_product_id = r[0]
                merge_product_name =r[1]
                order_add_time = r[2]
                order_price = r[3]
                order_id = r[4]
                order_pkid = r[5]
                product_id = r[6]
                product_name = r[7]
                source = "阿里云"

                order_list = (s,s,merge_product_id,merge_product_name,order_add_time,order_price,source,order_id,order_pkid,product_id,product_name)
                li.append(order_list)
            if(len(li)>200):
                sql21 = "insert into tj_standard_order(standard_company_name,account_company_name,merge_product_id,merge_product_name,order_add_time,order_pay,order_channel,order_id,order_pkid,product_id,product_name)\
                    values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                db.insertData(sql21,li)
                li = []
            else:
                pass
    sql4 = "insert into tj_standard_order(standard_company_name,account_company_name,merge_product_id,merge_product_name,order_add_time,order_pay,order_channel,order_id,order_pkid,product_id,product_name)\
        values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
    db.insertData(sql4,li)