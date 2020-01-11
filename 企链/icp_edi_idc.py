import json,datetime
import re
import pymysql
from DBUtils.PooledDB import PooledDB
import openpyxl


def transferContent(content):
    if content is None:
        return None
    else:
        string = ""
        for c in content:
            if c == '"':
                string += '\\\"'
            elif c == "'":
                string += "\\\'"
            elif c == "\\":
                string += "\\\\"
            else:
                string += c
        return string


def strB2Q(ustring):
    """把字符串全角转半角"""


def qingxi(ustring):
    ss = []
    for s in ustring:
        rstring = ""
        for uchar in s:
            inside_code = ord(uchar)
            if inside_code == 12288:  # 全角空格直接转换
                inside_code = 32
            elif (inside_code >= 65281 and inside_code <= 65374):  # 全角字符（除空格）根据关系转化
                inside_code -= 65248
            rstring += chr(inside_code)
        ss.append(rstring)
    return ''.join(ss)

MYSQL_HOST = '39.106.39.121'
USER = 'root'
PASSWORD = 'YUZ224102lss@#'
DB = 'ali'
PORT = 3306


class DBpool(object):
    def __init__(self):
        # 5为连接池里的最少连接
        self.pool = PooledDB(pymysql, 5, host=MYSQL_HOST,
                             user=USER, passwd=PASSWORD, db=DB, port=PORT)

    def insertData(self, sql, a):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.executemany(sql, a)
            conn.commit()
        except Exception as e:
            print(e)
        cursor.close()
        conn.close()

def returnStringDate(dat):
    if(dat !=None):
        return datetime.datetime.strftime(dat, "%Y-%m-%d")
    else:
        return dat

workbook = openpyxl.load_workbook(
    "C:\\Users\\YZY\\Documents\\gitwork\\企链\\改---影视审批.xlsx")

sheet = workbook.worksheets[0]

rows = sheet.max_row

aa = []
sql = "insert into certificate(companyName,licence,licence_no,service_category,end_date,industry_type,start_date) values(%s,%s,%s,%s,%s,%s,%s)"
db = DBpool()
for r in range(2, rows+1):
    companayName = sheet.cell(r, 2).value
    bianhao = sheet.cell(r, 1).value
    zhonglei = (sheet.cell(r, 3).value)
    # if(zhonglei.endswith('，')):
    #     zhonglei=zhonglei[0:-1]
    # print(zhonglei)
    youxiaoqi = sheet.cell(r, 4).value
    leixing = sheet.cell(r, 5).value
    hangye = sheet.cell(r, 6).value
    start_time = sheet.cell(r, 7).value

    youxiaoqi = returnStringDate(youxiaoqi)
    start_time = returnStringDate(start_time)

    companayName = qingxi(companayName)
    bianhao = qingxi(bianhao)
    #zhonglei = qingxi(zhonglei)
    #youxiaoqi = qingxi(youxiaoqi)
    leixing = qingxi(leixing)
    hangye = qingxi(hangye)

    aa.append((companayName, leixing, bianhao, zhonglei, youxiaoqi, hangye,start_time))

    if(len(aa) > 200):
        db.insertData(sql, aa)
        aa = []
db.insertData(sql, aa)
