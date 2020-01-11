import json,datetime
import re
import pymysql
from DBUtils.PooledDB import PooledDB
import openpyxl


def transferContent(content):
    if content is None:
        return None
    else:
        string = ""
        for c in content:
            if c == '"':
                string += '\\\"'
            elif c == "'":
                string += "\\\'"
            elif c == "\\":
                string += "\\\\"
            else:
                string += c
        return string


def strB2Q(ustring):
    """把字符串全角转半角"""


def qingxi(ustring):
    if(ustring!=None):
        ss = []
        for s in ustring:
            rstring = ""
            for uchar in s:
                inside_code = ord(uchar)
                if inside_code == 12288:  # 全角空格直接转换
                    inside_code = 32
                elif (inside_code >= 65281 and inside_code <= 65374):  # 全角字符（除空格）根据关系转化
                    inside_code -= 65248
                rstring += chr(inside_code)
            ss.append(rstring)
        return ''.join(ss)
    else:
        return ustring

MYSQL_HOST = '39.106.39.121'
USER = 'root'
PASSWORD = 'YUZ224102lss@#'
DB = 'ali'
PORT = 3306


class DBpool(object):
    def __init__(self):
        # 5为连接池里的最少连接
        self.pool = PooledDB(pymysql, 5, host=MYSQL_HOST,
                             user=USER, passwd=PASSWORD, db=DB, port=PORT)

    def insertData(self, sql, a):
        conn = self.pool.connection()
        cursor = conn.cursor()
        try:
            cursor.executemany(sql, a)
            conn.commit()
        except Exception as e:
            print(e)
        cursor.close()
        conn.close()

def returnStringDate(dat):
    if(dat !=None):
        return datetime.datetime.strftime(dat, "%Y-%m-%d")
    else:
        return dat

workbook = openpyxl.load_workbook(
    "C:\\Users\\YZY\\Documents\\gitwork\\企链\\游戏出版及运营.xlsx")

sheet = workbook.worksheets[0]

rows = sheet.max_row

aa = []
sql = "insert into game_license(game_name,declaration_category,publishing_unit,operation_unit,symbol,publishing_number,publish_time,publishing_record,operation_record,industry_type) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
db = DBpool()
for r in range(2, rows+1):
    game_name = sheet.cell(r, 1).value
    declaration_category = sheet.cell(r, 2).value
    publishing_unit = sheet.cell(r, 3).value
    operation_unit = sheet.cell(r, 4).value
    symbol = sheet.cell(r, 5).value
    publishing_number = sheet.cell(r, 6).value
    publish_time = (sheet.cell(r, 7).value)
    print(type(publish_time))
    publishing_record = sheet.cell(r, 9).value
    operation_record = sheet.cell(r, 8).value
    industry_type = sheet.cell(r, 10).value

    game_name = qingxi(game_name)
    declaration_category = qingxi(declaration_category)
    publishing_unit = qingxi(publishing_unit)
    operation_unit = qingxi(operation_unit)
    symbol = qingxi(symbol)
    publishing_number = qingxi(publishing_number)
    #publish_time = qingxi(publish_time)
    publishing_record = qingxi(publishing_record)
    operation_record = qingxi(operation_record)
    industry_type = qingxi(industry_type)


    aa.append((game_name,declaration_category,publishing_unit,operation_unit,symbol,publishing_number,publish_time,publishing_record,operation_record,industry_type))

    if(len(aa) > 200):
        db.insertData(sql, aa)
        aa = []
db.insertData(sql, aa)




